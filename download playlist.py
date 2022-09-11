import googleapiclient.discovery
from urllib.parse import parse_qs, urlparse
from pytube import YouTube
import pandas as pd
from openpyxl import load_workbook


#extract playlist id from url
url = 'https://www.youtube.com/playlist?list=PLBnvQD5J8X_XVY-Y4oI0tMky0ahWN5stj'
query = parse_qs(urlparse(url).query, keep_blank_values=True)
playlist_id = query["list"][0]

print(f'get all playlist items links from {playlist_id}')
youtube = googleapiclient.discovery.build("youtube", "v3", developerKey = "apikey")

request = youtube.playlistItems().list(
    part = "snippet",
    playlistId = playlist_id,
    maxResults = 420
)
response = request.execute()

playlist_items = []
while request is not None:
    response = request.execute()
    playlist_items += response["items"]
    request = youtube.playlistItems().list_next(request, response)

print(f"total: {len(playlist_items)}")
link=([
    f'https://www.youtube.com/watch?v={t["snippet"]["resourceId"]["videoId"]}&list={playlist_id}&t=0s'
    for t in playlist_items
])
print(link)

path = executable_path=r"C:\Users\KDK\Desktop\ytbe_dnwload"

y=2
x=0
while y < 409:  ## length playlistu (407) + variable 2
    df = pd.read_excel(r"C:\Users\KDK\Desktop\ytbe_dnwload\links.xlsx")
    wb = load_workbook(r"C:\Users\KDK\Desktop\ytbe_dnwload\links.xlsx")
    ws = wb.worksheets[0]

    ws.cell(row=y, column=3).value = link[x]
    x+=1
    y += 1
    wb.save(r"C:\Users\KDK\Desktop\ytbe_dnwload\links.xlsx")
    print(y)


